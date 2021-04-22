#@author :SHUBHAM SHARMA
#This is a demonstration tool used for demonstrating Calibration of Point Targets 
#In SAR imagery

from scipy.interpolate import griddata,interp1d
from mpldatacursor import datacursor
from osgeo import gdal
import matplotlib.pyplot as plt
import numpy as np
from tkinter import *
from tkinter import ttk
import tkinter as tk
from tkinter.filedialog import askopenfilename
import os
from PIL import ImageTk,Image
from tkinter import messagebox
import math
from numpy import unravel_index 
import csv
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import scipy
from scipy.interpolate import spline
from scipy.signal import argrelextrema
from scipy.ndimage.filters import convolve


#Too many global variables have been utilised here as they were used for various sensors which can be avoided when using only one sensor

global b3,b4,t1,m1,n1,ctr,Ex,wv,bmcc,d,e,s,t,m,n,gm,inc,sen,ccb,win,inc_map,lut_g,sigma_ref,ph,s7,t7,lin_spac,pix_spac
global a_arr,b_arr,c_arr,d_arr,a1,b1,c1,d1,e1,a_garr,b_garr,c_garr,d_garr,e_garr,r_b,c_b,inc_n,beta,r_sigma,r_gamma,check1,check2

pt=list()
ga=list()
in_cal=list()

root = tk.Tk()
sc_w=root.winfo_screenwidth()
sc_h=root.winfo_screenheight()
sc_w=int(sc_w/2)
sc_h=int(sc_h/3)


root.geometry('{}x{}'.format(sc_w,sc_h))


class GISApplication(object):
    
    def __init__(self):
        self.filename = None
        self.outfilename = None
                
        
    def open_file(self):
        self.filename  = askopenfilename(initialdir="C:/Users/admin 1/.spyder2-py3",
                           filetypes =(("All Files","*.*"),("Text File", "*.txt")),
                           title = "Choose a file."
                           )

        return self.filename 
         
    def get_file(self):
        if self.filename != None:
            return self.filename
        else:
            return self.open_file()
        
        
application = GISApplication()                           
tm= GISApplication()  
fm= GISApplication()                          
hm=GISApplication()
am=GISApplication()
mm=GISApplication()
im=GISApplication()
rm=GISApplication()
jm=GISApplication()
lut=GISApplication()
tra=GISApplication()

Title = root.title( "Calibration Toolbox")



label = ttk.Label(root, text ="Sentinel-1 SAR CALIBRATION TOOLBOX\n\n",foreground="Dark Blue",font=("Times New Roman", 20))


label.grid(row=0,column=0,sticky=N+S+W,padx=(round(sc_w/3.5)),pady=0)



lbl = ttk.Label(root, text ="INPUT",foreground="Blue",font=("Times New Roman", 15))

lbl.grid(row=1,column=0,sticky=N+S+W,padx=sc_h,pady=20)


#Menu Bar
menu = Menu(root)
root.config(menu=menu)

file = Menu(menu)



def onclick_input(event=None):
    

    global t1,pt,sen,pg
    t1=tm.get_file()
    

    #Histogram equalization for better display of Sentinel-1 image
    def histeq(im):
        global sen
        
        if (sen=='Sentinel-1'): 
            
            nbins=65536
        elif  sen=='RISAT-1':
            nbins=256
        elif sen=='Radarsat-1':
            nbins=256
        elif sen=='ALOS-2 PALSAR':
            nbins=256
            
        
        
        imhist,bins=np.histogram(im.flatten(),nbins,normed=True)
        imhist[0]=0
        cdf=imhist.cumsum()
        cdf**.5
        cdf=(nbins-1)*cdf/cdf[-1]
        im2=np.interp(im.flatten(),bins[:-1],cdf)
        return np.array(im2,int).reshape(im.shape)
        
        
    
    
    
    
    
    
    #onclick event   
    def oc1(event):
        global x1,y1,m1,n1,ctr,pt
        if event.dblclick :

            x1=event.xdata
            y1=event.ydata
            x1=int(round(x1))
            y1=int(round(y1))
            pt.append((x1,y1))
            #print(m1)   
            if len(pt)==5:
                fig.canvas.mpl_disconnect(cid)
                plt.close(1)
                #print(pt)
            
        return   
    
    
    #Open Sentinel-1 Image for displaying
    if (sen=='Sentinel-1') and '.tiff' in t1:
        
        
        x=gdal.Open(t1)
        
        if  ((x.ReadAsArray()).dtype=='complex64') or ((x.ReadAsArray()).dtype=='complex128'):
            #For Single Look Complex Image
            
             x=x.ReadAsArray() 
             x_re=np.real(x)
             x_im=np.imag(x)
            
            
             fig,ax=plt.subplots()
             power=x_re**2+x_im**2
             power=histeq(power)
    
             fig.canvas.set_window_title('Power Image ')
             myax=ax.imshow(power,cmap='gray',interpolation='nearest')
             datacursor(myax,hover=True,bbox=dict(alpha=1,fc='w'))
             ax.set_title("Power Image")
             
             cid=fig.canvas.mpl_connect('button_press_event',oc1)
 

   
        else:
            x=x.ReadAsArray() 
            
            fig,ax=plt.subplots()
            power=x**2
            power=histeq(power)
    
            fig.canvas.set_window_title('Power Image ')
            myax=ax.imshow(power,cmap='gray',interpolation='nearest')
            datacursor(myax,hover=True,bbox=dict(alpha=1,fc='w'))
            ax.set_title("Power Image")
        
            cid=fig.canvas.mpl_connect('button_press_event',oc1)
            
             
             
    else :
        messagebox.showinfo(title="Warning!!!!!",message="File Format not Supported")
            
            
            
        
        
        
    return    
    
    
    



b2 = Button(root,text="Input File", command=onclick_input)

b2.grid(row=10,column=0,sticky='W',padx=round(sc_w/3.2))

OPTIONS = ["--Sensor--",
"Sentinel-1",
] 


master=root


variable = StringVar(master)
variable.set(OPTIONS[0]) # default value

w = OptionMenu(master, variable, *OPTIONS)
w.grid(row=10,column=0,sticky='W',padx=(sc_w/8))



def ok():
    global sen,Ex,wv
    print ("value is:" + variable.get())
    sen=variable.get()
    
            
    if (sen=='Sentinel-1') :
        
        wv=0.0555   #wavelength is 0.0555 m for Sentinel-1
        
        c11="1.Please Select Input file.\n\n"
        c12="2.After Selecting Input Image,select 5 points,one for Corner reflector\n"
        c13="and four for clutter centres on the image.\n\n"
        c14="3. After Input File Selection, Select Calibration.xml File option\n" 
        c15="to enter Calibration.xml file.\n\n"
        c16="4.Proceed to OUTPUT options.\n"
        messagebox.showinfo(title="Sentinel-1 Guide",message=c11+c12+c13+c14+c15+c16)
        
   
         
         

button = Button(master, text="OK", command=ok)
button.grid(row=11,column=0,sticky='W',padx=round(sc_w/6.154))



#For calibration to Sigma naught and gamma naught
def calibration(event=None):
    
    global x1,y1,m1,n1,ctr,pt,ga,a_arr,b_arr,c_arr,d_arr,e_arr,a1,b1,c1,d1,e1
    global a_garr,b_garr,c_garr,d_garr,e_garr
    file=am.get_file()
    

    def calib(m,n):
        
        
        
        pp=list()
        ls_sigma=list()
        ls_gamma=list()
        from xml.dom import minidom
        xmldoc = minidom.parse(file)
        itemlist = xmldoc.getElementsByTagName('line') 
        sigma_n=xmldoc.getElementsByTagName('sigmaNought') 
        gamma_n=xmldoc.getElementsByTagName('gamma') 
        

        for i in range(len(itemlist)):
            pp.append(int(itemlist[i].firstChild.nodeValue))
            
            
        for i in range(len(pp)):
            if pp[i]<=m<=pp[i+1]:
                index=i
            
        tt1="".join(sigma_n[index].firstChild.nodeValue.split())
        tt2="".join(gamma_n[index].firstChild.nodeValue.split())
            
            
        for i in range(0,len(tt1),12):
            ls_sigma.append(float(tt1[i:i+12]))
            ls_gamma.append(float(tt2[i:i+12]))   
            
        
        index_sigma=int(n/40)
        index_unravel=n%40

        sigma_arr=ls_sigma[index_sigma-1:index_sigma+3]
        gamma_arr=ls_gamma[index_sigma-1:index_sigma+3]
        arr1=np.linspace(sigma_arr[0],sigma_arr[1],num=40,endpoint=False)
        arr2=np.linspace(sigma_arr[1],sigma_arr[2],num=40,endpoint=False)
        arr3=np.linspace(sigma_arr[2],sigma_arr[3],num=40,endpoint=True)
        ar=np.concatenate((arr1,arr2,arr3),axis=0)
        arr1=np.linspace(gamma_arr[0],gamma_arr[1],num=40,endpoint=False)
        arr2=np.linspace(gamma_arr[1],gamma_arr[2],num=40,endpoint=False)
        arr3=np.linspace(gamma_arr[2],gamma_arr[3],num=40,endpoint=True)
        ar1=np.concatenate((arr1,arr2,arr3),axis=0)
        print(ar1)
        print(sigma_arr)
        ind_pos=40+index_unravel
        print(ind_pos)
        return ar,ind_pos,ar1

    a_arr,a1,a_garr=calib(pt[0][1],pt[0][0])
    b_arr,b1,b_garr=calib(pt[1][1],pt[1][0])
    c_arr,c1,c_garr=calib(pt[2][1],pt[2][0])
    d_arr,d1,d_garr=calib(pt[3][1],pt[3][0])
    e_arr,e1,e_garr=calib(pt[4][1],pt[4][0])
    
    
     
b22 = Button(root,text="Calibration.xml File", command=calibration)
print('\n\n\n')
b22.grid(row=10,column=0,sticky='W',padx=(sc_w/1.25),columnspan=2)
 
 

# Annotation.xml File
def annotation(event=None):
    
    global gm,d,e
    
    fi2=im.get_file()
    from xml.dom import minidom
    xmldoc = minidom.parse(fi2)
    itemlist = xmldoc.getElementsByTagName('line') 
    gm=xmldoc.getElementsByTagName('incidenceAngleMidSwath') 
    d=xmldoc.getElementsByTagName('rangePixelSpacing') 
    e=xmldoc.getElementsByTagName('azimuthPixelSpacing') 


    gm=math.radians(float(gm[0].firstChild.nodeValue))
    d=float(d[0].firstChild.nodeValue)
    e=float(e[0].firstChild.nodeValue)
    
   
    

b99 = Button(root,text="Annotation.xml File", command=annotation)
print('\n\n\n')
b99.grid(row=10,column=0,sticky='W',padx=(sc_w/1.9),columnspan=2)




lbl1 = ttk.Label(root, text ="\n\nOUTPUT",foreground="Blue",font=("Times New Roman", 15))
lbl1.grid(row=400,column=0,sticky=N+S+W,padx=sc_h,pady=20)    
    
    


#Radiometric Calibration
def onclick_radcal(event=None):
    
    global inc
    top = Toplevel()
    top.geometry('{}x{}'.format(int(sc_w/4), int(sc_h/2)))
    titl=top.title("Radiometric Calibration")
    
    def point(event=None):
        
        global wv,d,e,gm,bm_cc,inc,sen,a_arr,b_arr,c_arr,d_arr,e_arr,a1,b1,c1,d1,e1,m1,n1,inc_map,lut_s,sigma_ref,r_sigma,lin_spac,pix_spac
        g=gdal.Open(t1)
        
        
        if (sen=='Sentinel-1') and'.tiff' in t1:
            
            
            if  ((g.ReadAsArray()).dtype=='complex64') or ((g.ReadAsArray()).dtype=='complex128'): 
                
                
                gi=g.ReadAsArray()
                x_re=np.real(gi)
                x_im=np.imag(gi)
                power=x_re**2+x_im**2
            
            else:
                gi=g.ReadAsArray()
                power=gi**2
                        
            
            
        else:
            print("Invalid Sensor!!! ")
        
        
        
        
        
        tip=Toplevel()
        tip.geometry('{}x{}'.format(int(sc_w/1.6), sc_h))
        
        tit1=tip.title("Point Target Details")
        
        
        
        L1=Label(tip,text="Point Target Window size Line (Row)")
        L1.grid(row=20,column=10)
        E1=Entry(tip,bd=5)
        E1.grid(row=20,column=20)
        
        
        L2=Label(tip,text="Point Target Window size Pixel(Column)")
        L2.grid(row=30,column=10)
        E2=Entry(tip,bd=5)
        E2.grid(row=30,column=20)
        
        L3=Label(tip,text=" Clutter Window Size")
        L3.grid(row=40,column=10)
        E3=Entry(tip,bd=5)
        E3.grid(row=40,column=20)
        
        
        
        L4=Label(tip,text="Corner Reflector Type")
        L4.grid(row=50,column=10)
        E4=Entry(tip,bd=5)
        E4.grid(row=50,column=20)
        
        L5=Label(tip,text="Leg Length (in m)")
        L5.grid(row=60,column=10)
        E5=Entry(tip,bd=5)
        E5.grid(row=60,column=20)
        
        
        
        def onclick_incidence(event=None):
            
         
            
            
            
            def oc(event):
                global inc
                inc=float(Ex1.get())
                print(inc)
        
        Lx1=Label(tip,text="Incidence Angle at Point target(in deg.)")
        Lx1.grid(row=80,column=10)
        Ex1=Entry(tip,bd=5)
        Ex1.grid(row=80,column=20)
        Ex1.bind('<ButtonPress>',onclick_incidence)   
    
    
    
        L8=Label(tip,text="File Name (.xlsx)")
        L8.grid(row=90,column=10)
        E8=Entry(tip,bd=5)
        E8.grid(row=90,column=20)
        
        
        L9=Label(tip,text="C R Name")
        L9.grid(row=100,column=10)
        E9=Entry(tip,bd=5)
        E9.grid(row=100,column=20)
        
        
        if (sen=='Sentinel-1'):
            Ex1.config(state='disabled')
            
        
        
        s1="In Point Target window Size, values must be integer and must be odd\n"
        s2="In Clutter Window size window should be square\n"
        s3="Distance between pixels should be in integers\n"
        s4="In corner Reflector Type, Enter :\n  1 for Triangular Trihedral\n"
        s5="2 for Flat Square Plate\n 3 for Circular Trihedral\n"
        s6="4 for Dihedral\n 5 for Square Trihedral\n Press OK only after entering all the values\n\n"
        s7="If Incidence Angle File is not available, enter the Incidence angle in degrees\n"
        s8="in the given entry box and Double Click the box."
        s9="After entering incidence Angle in the entry box, Double click the inside the box"
        messagebox.showinfo(title="Information",message=s1+s2+s3+s4+s5+s6+s7+s8+s9)
        
        row1=int(E1.get())
        col1=int(E2.get())
        cwsize=int(E3.get())
        
    
        if sen=='Sentinel-1':
            
            
        
            row=math.floor(row1/2)
            col=math.floor(col1/2) 
        
            n1=pt[0][1]
            m1=pt[0][0]
            
            pw=power[n1-row:n1+row+1, m1-col:m1+col+1]
            ti=(n1-row,m1-row)
            

            hi=unravel_index(pw.argmax(),pw.shape)
            tt=(ti[0]+hi[0],ti[1]+hi[1])
            
            n1=tt[0]
            m1=tt[1]
            print(n1,m1)
            
            power_pixel=power[n1-row:n1+row+1, m1-col:m1+col+1]
            print(power_pixel)
            arr=np.array([a_arr,]*row1)
           
            arr=arr[0:row1,a1-col:a1+col+1]
            
            print("\n arr is\n")
            print(arr)
            power_pixel=power_pixel/(arr**2)
            print("Power Pixel is :")
            print(power_pixel)
            print(power_pixel.shape)

            ki,k=cwsize,cwsize
            
            #here cl1,cl2,cl3,cl4 represent 4 clutter centres for point calibration
            if (ki % 2) !=0 :
                
                k=math.floor(ki/2)
                
                cl1=power[pt[1][1]-k:pt[1][1]+k+1,pt[1][0]-k:pt[1][0]+k+1]
                arr=np.array([b_arr,]*ki)
          
                arr=arr[0:ki,b1-k:b1+k+1]
                print(arr)
                cl1=cl1/(arr**2)

                cl2=power[pt[2][1]-k:pt[2][1]+k+1,pt[2][0]-k:pt[2][0]+k+1]   
                arr=np.array([c_arr,]*ki)
           
                arr=arr[0:ki,c1-k:c1+k+1]
                print(arr)
                cl2=cl2/(arr**2)
                
                cl3=power[pt[3][1]-k:pt[3][1]+k+1,pt[3][0]-k:pt[3][0]+k+1] 
                arr=np.array([d_arr,]*ki)
             
                arr=arr[0:ki,d1-k:d1+k+1]
                print(arr) 
                cl3=cl3/(arr**2)

                cl4=power[pt[4][1]-k:pt[4][1]+k+1,pt[4][0]-k:pt[4][0]+k+1]
                arr=np.array([e_arr,]*ki)
              
                arr=arr[0:ki,e1-k:e1+k+1]
                print(arr)
                cl4=cl4/(arr**2)
            
            
               


            else:
                
                
                k=ki-1
            
                cl1=power[pt[1][1]-k:pt[1][1]+1,pt[1][0]-k:pt[1][0]+1]
                arr=np.array([b_arr,]*ki)
              
                arr=arr[0:ki,b1-k:b1+1]
                print(arr)
                cl1=cl1/(arr**2)
                
                cl2=power[pt[2][1]-k:pt[2][1]+1,pt[2][0]-k:pt[2][0]+1]
                arr=np.array([c_arr,]*ki)
             
                arr=arr[0:ki,c1-k:c1+1]
                print(arr)
                cl2=cl2/(arr**2)

        
                cl3=power[pt[3][1]-k:pt[3][1]+1,pt[3][0]-k:pt[3][0]+1]
                arr=np.array([d_arr,]*ki)
              
                arr=arr[0:ki,d1-k:d1+1]
                print(arr)
                cl3=cl3/(arr**2)

                cl4=power[pt[4][1]-k:pt[4][1]+1,pt[4][0]-k:pt[4][0]+1]   
                arr=np.array([e_arr,]*ki)
               
                arr=arr[0:ki,e1-k:e1+1]
                print(arr)
                cl4=cl4/(arr**2)


        pixel_sum=power_pixel.sum()
       
        print(pixel_sum)
        clut_pixel=(cl1.sum()+cl2.sum()+cl3.sum()+cl4.sum())/(4*cwsize*cwsize)
        bp=clut_pixel
        bip=10*math.log10(bp)
        print("Eu %f in db"%(bip))
        cl_power=clut_pixel*(row1*col1)
        back_corr_power=pixel_sum-cl_power
        bcp=back_corr_power
        back_corr_power=10*math.log10(back_corr_power)
        
        print("background Power %f"%(bcp))
        print("bcp db%f"%(back_corr_power))
        cl_power=10*math.log10(cl_power)
        
        a=int(E4.get())
        b=float(E5.get())
        
        c=wv
        scr=bcp/bp
        scr=10*math.log10(scr)
        
        
        
        
        hi=E8.get()
        ii=E9.get()
        hi=hi+'.xlsx'
        
        
        if a==1:
            sigma_ref=(4*math.pi*(b**4))/(3*c*c)
            crt="Triangular Trihedral"
        elif a==2:
            sigma_ref=(4*math.pi*(b**4))/(c*c)
            crt="Flat Square Plate"
        elif a==3:
            sigma_ref=(0.507*(math.pi**3)*(b**4))/(c*c)
            crt="Circular Trihedral"
        elif a==4:
            sigma_ref=(8*math.pi*(b**4))/(c*c)
            crt="Dihedral"
        elif a==5:
            sigma_ref=(12*math.pi*(b**4))/(c*c)
            crt="Square Trihedral"
        
 
        if sen=='Sentinel-1':
            
            
            
            sigma_ref=10*math.log10(sigma_ref)
            area=d*e
            
            if ((g.ReadAsArray()).dtype=='complex64') or ((g.ReadAsArray()).dtype=='complex128'):
                 
                rcs=bcp*(area/math.sin(gm))
            else:
                 rcs=bcp*area
            
           
            
            
            rcs=10*math.log10(rcs)
            diff=rcs-sigma_ref
            print(diff,rcs,sigma_ref)
            tip1=Toplevel(top)
            tip1.geometry('{}x{}'.format(400,300))
            
            si="Background Corrected Sigma Nought : {:0.2f} dB".format(back_corr_power)
            ti="\nClutter Sigma Nought : {:0.2f} dB".format(cl_power)
            ui="\nTheoretical RCS : {:0.2f} dBm^2".format(sigma_ref)
            vi="\nSignal to Clutter Ratio : {:0.2f} dB".format(scr)
            di="\n Calculated RCS {:0.2f} dBm^2".format(rcs)
            ei="\n Difference in RCS {:0.2f} dBm^2".format(diff)
            tex=Text(tip1,height=5,width=50)
            tex.grid()
            tex.insert(END,si+ti+ui+di+ei+vi)
            tex.mainloop()
            tip1.mainloop()
            
    
            if os.path.isfile(hi)==False:
                
                filepath = hi
                wb = openpyxl.Workbook()
                
            
            
         
                wb.save(filepath)
                df1 = pd.DataFrame({'Back. Corr. Sigma Nought(dB)': [back_corr_power],'SCR (dB)':[scr],'Theoretical RCS (dBm^2)':[sigma_ref],'Calculated RCS (dbm^2)':[rcs],'Difference in RCS (dbm^2)':[diff],'CR Type':[crt],'CR No.':[ii]})
                book = load_workbook(filepath)
                writer = pd.ExcelWriter(filepath, engine = 'openpyxl')
                writer.book = book
                df1.to_excel(writer, sheet_name = ii,index=False)
                writer.save()
                writer.close()
        
            
            else:
                
                df1 = pd.DataFrame({'Back. Corr. Sigma Nought(dB)': [back_corr_power],'SCR (dB)':[scr],'Theoretical RCS (dBm^2)':[sigma_ref],'Calculated RCS (dbm^2)':[rcs],'Difference in RCS (dbm^2)':[diff],'CR Type':[crt],'CR No.':[ii]})
                path = hi
 
                book = load_workbook(path)
                writer = pd.ExcelWriter(path, engine = 'openpyxl')
                writer.book = book
                df1.to_excel(writer, sheet_name = ii,index=False)
       
                writer.save()
            
        
        
#Distributed Target  Calibration      
    def dist(event=None):
        global gm,sen,a_arr,b_arr,c_arr,d_arr,a1,b1,c1,d1,e1,a_garr,b_garr,c_garr,d_garr,e_garr,t1,pt,win,ccb
        global inc1,inc2,inc3,inc4,inc5,in_cal,pt,lut_s,lut_g
        
        if sen=='Sentinel-1':
            
            tip1=Toplevel(top)
            tip1.geometry('{}x{}'.format(400,200))
            
            L12=Label(tip1,text="Enter window Size")
            L12.grid(row=90,column=10)
            E12=Entry(tip1,bd=5)
            E12.grid(row=90,column=20)
            s1="1.Enter window size in no. of pixels\n\n"
            s2="2.The window size must be odd.\n\n"
            messagebox.showinfo(title="Information",message=s1+s2)
            ki=int(E12.get())
            k=math.floor(ki/2)
            print(ki)
              
            x=(gdal.Open(t1)).ReadAsArray()
            x_re=np.real(x)
            x_im=np.imag(x)
            power=x_re**2+x_im**2
            
            
            def do_cal(k,main,arr,ki,d1):
                
                arr_l=np.array([arr,]*ki)
                arr_l=arr_l[0:ki,d1-k:d1+k+1]
                
                print(arr_l)
                
                main=main/(arr_l**2)
                
                return main
                
                
                
                
            
            main=power[pt[0][1]-k:pt[0][1]+k+1,pt[0][0]-k:pt[0][0]+k+1]
            print(main)
            main=do_cal(k,main,a_arr,ki,a1)
            main_g=do_cal(k,main,a_garr,ki,a1)
            
            cl1=power[pt[1][1]-k:pt[1][1]+k+1,pt[1][0]-k:pt[1][0]+k+1]
            cl1=do_cal(k,cl1,b_arr,ki,b1)
            cl1_g=do_cal(k,cl1,b_garr,ki,b1)
            
            cl2=power[pt[2][1]-k:pt[2][1]+k+1,pt[2][0]-k:pt[2][0]+k+1] 
            cl2=do_cal(k,cl2,c_arr,ki,c1)  
            cl2_g=do_cal(k,cl2,c_garr,ki,c1)  
            
            cl3=power[pt[3][1]-k:pt[3][1]+k+1,pt[3][0]-k:pt[3][0]+k+1]  
            cl3=do_cal(k,cl3,d_arr,ki,d1)  
            cl3_g=do_cal(k,cl3,d_garr,ki,d1)
            
            cl4=power[pt[4][1]-k:pt[4][1]+k+1,pt[4][0]-k:pt[4][0]+k+1]   
            cl4=do_cal(k,cl4,e_arr,ki,e1)  
            cl4_g=do_cal(k,cl4,e_garr,ki,e1)
            
            main_m=10*math.log10(np.mean(main))
            cl1_m=10*math.log10(np.mean(cl1))
            cl2_m=10*math.log10(np.mean(cl2))
            cl3_m=10*math.log10(np.mean(cl3))
            cl4_m=10*math.log10(np.mean(cl4))
            
            main_gm=10*math.log10(np.mean(main_g))
            cl1_gm=10*math.log10(np.mean(cl1_g))
            cl2_gm=10*math.log10(np.mean(cl2_g))
            cl3_gm=10*math.log10(np.mean(cl3_g))
            cl4_gm=10*math.log10(np.mean(cl4_g))
            
            tip2=Toplevel(tip1)
            tip2.geometry('{}x{}'.format(600,200))
            
            si="Mean Sigma Nought around First Selected point: {:0.2f} dB\n".format(main_m)
            ti="\nMean Gamma Nought around First Selected point: {:0.2f} dB\n\n".format(main_gm)
            
            ui="Mean Sigma Nought around Second Selected point: {:0.2f} dB\n".format(cl1_m)
            vi="\nMean Gamma Nought around Second Selected point: {:0.2f} dB\n\n".format(cl1_gm)
            
            di="Mean Sigma Nought around Third Selected point: {:0.2f} dB\n".format(cl2_m)
            li="Mean Gamma Nought around Third Selected point: {:0.2f} dB\n\n".format(cl2_gm)
            
            mi="Mean Sigma Nought around Fourth Selected point: {:0.2f} dB\n".format(cl3_m)
            ghi="Mean Gamma Nought around Fourth Selected point: {:0.2f} dB\n\n".format(cl3_gm)
            
            ri="Mean Sigma Nought around Fifth Selected point: {:0.2f} dB\n".format(cl4_m)
            ji="Mean Gamma Nought around Fifth Selected point: {:0.2f} dB\n\n".format(cl4_gm)
            
            tex=Text(tip2,height=50,width=100)
            tex.grid()
            tex.insert(END,si+ti+ui+vi+di+li+mi+ghi+ri+ji)
            tex.mainloop()
            tip2.mainloop()
            
            tip1.mainloop()
            
            
    br1=Button(top,text="Point Target", command=point)
    br1.grid(row=5,column=0,sticky=W,padx=50,pady=30)
    
    
    br2=Button(top,text="Distributed Target", command=dist)
    br2.grid(row=8,column=0,sticky=W,padx=50,pady=30)
    
    top.mainloop()
     
b6 = Button(root,text="Radiometric Calibration", command=onclick_radcal)
print('\n\n\n')
b6.grid(row=900,column=0,sticky='W',padx=round(sc_w/5.01)) 



#SAR Quality Parameters
def sar_qual(event=None):
    #Calculates Peak Side Lobe Ratio and Integrated Side Lobe Ratio parameters
    #to assess Image quality
    global n1,m1,inc_n,gm,sigma_ref,a_arr,d,e,a1,lut_s,ccb,check1,check2,r_gamma,r_sigma,lin_spac,pix_spac
    
    sar_top = Toplevel()
    sar_top.geometry('{}x{}'.format(int(sc_w/4), int(sc_h/2)))
    
    
    if sen=='Sentinel-1':   
        
        
        x=(gdal.Open(t1)).ReadAsArray()
        x_re=np.real(x)
        x_im=np.imag(x)
        power=x_re**2+x_im**2
          
        xm=np.linspace(-5,4,num=10,endpoint=True,retstep=True)
        ym=np.linspace(-4,5,num=10,endpoint=True,retstep=True)

        xm,ym=np.meshgrid(xm[0],ym[0])


        xi=np.linspace(-5,4,num=46,endpoint=True,retstep=True)
        yi=np.linspace(-4,5,num=46,endpoint=True,retstep=True)
    
        X, Y = np.meshgrid(xi[0],yi[0])
        print(n1,m1)
     
            
        zm=power[n1-5:n1+5,m1-5:m1+5]
        gain_sig=np.array([a_arr,]*10)
        gain_sig=gain_sig[0:10,a1-5:a1+5]
        zm=zm/(gain_sig)
        zm=10*np.log10(zm)
        print(zm)
        
        zmax=np.nanmax(zm)
        zm=zm-zmax


        

        Z=griddata((xm.flatten(),ym.flatten()),zm.flatten(),(X,Y),method='cubic')
        fig = plt.figure(1)
        ax = fig.gca(projection='3d')
        ax.set_xlabel("Range")
        ax.set_ylabel("Azimuth")
        ax.set_zlabel("Relative Power (in dB)")

        surf = ax.plot_surface(X, Y, Z,rstride=1,cstride=1, cmap='rainbow',linewidth=0, antialiased=False)
        fig.colorbar(surf, shrink=0.5, aspect=5)
        plt.show()        
            
            
            
            
            
        y=power[n1,m1-8:m1+9]
        gain_sig=np.array([a_arr,]*1)
        gain_sig=gain_sig[0,a1-8:a1+9]
        y=y/(gain_sig**2)
        
        
        ra_islr=y
        
        x_is=np.linspace(-8,8,num=17,endpoint=True,retstep=True)
        xnew_islr=np.linspace(-8,8,300)
        
        
        power_tm_islr=spline(x_is[0],ra_islr,xnew_islr)
        
        
        minima_islr=argrelextrema(power_tm_islr, np.less)
        maxima_islr=argrelextrema(power_tm_islr, np.greater)
        
        
        maximum=power_tm_islr[maxima_islr]
        
        pos_max=int(np.argwhere(maximum==np.max(maximum)))
        
        pos_max2=int(maxima_islr[0][pos_max])
        
        
        for i in range(len(minima_islr[0])):
            
            if pos_max2 in range(minima_islr[0][i],minima_islr[0][i+1]) :
                
                sl1=minima_islr[0][i]
                sl2=minima_islr[0][i+1]
                break
                 
             
            
                
        
        power_sl1=power_tm_islr[0:sl1+1]
        power_sl1=sum(power_sl1[power_sl1>0])
         
        power_sl2=sum(power_tm_islr[sl2:])
        power_sl2=sum(power_sl2[power_sl2>0])
         
        power_ml=sum(power_tm_islr[sl1+1:sl2])
        islr=(power_sl1+power_sl2)/power_ml
        print(power_sl1,power_sl2,power_ml,islr)
        islr_range=10*np.log10(islr)
        
        
        
        
        y=10*np.log10(y)
        
        
        
        mini=min(filter(lambda x:x!=float('-inf'),y))
        y[np.isnan(y)]=mini
        y[np.isinf(y)]=mini
        ymax=np.nanmax(y)
        y=y-ymax
        x0=np.linspace(-8,8,num=17,endpoint=True,retstep=True)
        xnew=np.linspace(-8,8,300)
        power_sm=spline(x0[0],y,xnew)
        
        minima=argrelextrema(power_sm, np.less)
        maxima=argrelextrema(power_sm, np.greater)
        
        
        mlsl=power_sm[maxima]
        
        
        Is=np.max(mlsl[mlsl<=0])
        
        pslr_r=Is
    
    
        maxi=power_sm[maxima]
        maxi2=maxima[0]
    
        aa=int(np.argwhere(maxi==np.max(maxi)))
        
        pos = (np.abs(power_sm-(-3))).argmin()
        a11=np.argsort(np.abs(power_sm[0:maxi2[aa]]-(-3)))[0]
        a22=np.argsort(np.abs(power_sm[maxi2[aa]:]-(-3)))[0]
        a22=maxi2[aa]+a22
    
        pix_space=e
        resolution_r=(abs(xnew[a11])+abs(xnew[a22]))*pix_space
    
    
        fig, ax = plt.subplots(1, 2, figsize=(8, 5), sharex=True, sharey=False,
                       subplot_kw={'adjustable': 'box-forced'})

        ax0,ax1= ax.ravel()
        plt.tight_layout()


        ax0.plot(x0[0],y,'go',xnew,power_sm,'b-',linewidth=2.0)

        ax0.set_xlabel('Range (m)',fontsize=20)
        ax0.set_ylabel('Relative Power (dB)',fontsize=20)
        ax0.tick_params(axis='x',labelsize=20)
        ax0.tick_params(axis='y',labelsize=20)

        ax0.axis('on')

            
       # Azimuth
        t=power[n1-8:n1+9,m1]

        gain_sig=np.array([a_arr,]*17)
        gain_sig=gain_sig[:,a1]
        t=t/(gain_sig**2)
        
        az_islr=t
        
        x_is=np.linspace(-8,8,num=17,endpoint=True,retstep=True)
        xnew_islr=np.linspace(-8,8,300)
        
        
        power_tm_islr=spline(x_is[0],az_islr,xnew_islr)
        
        minima_islr=argrelextrema(power_tm_islr, np.less)
        maxima_islr=argrelextrema(power_tm_islr, np.greater)
        
        
        maximum=power_tm_islr[maxima_islr]
        
        pos_max=int(np.argwhere(maximum==np.max(maximum)))
        
        pos_max2=int(maxima_islr[0][pos_max])
        
        
        for i in range(len(minima_islr[0])):
            
            if pos_max2 in range(minima_islr[0][i],minima_islr[0][i+1]) :
                
                sl1=minima_islr[0][i]
                sl2=minima_islr[0][i+1]
                break
                 
             
            
                
        
        power_sl1=power_tm_islr[0:sl1+1]
        power_sl1=sum(power_sl1[power_sl1>0])
         
        power_sl2=power_tm_islr[sl2:]
        power_sl2=sum(power_sl2[power_sl2>0])

        power_ml=sum(power_tm_islr[sl1+1:sl2])
         
        islr=(power_sl1+power_sl2)/power_ml
        print(islr)
        islr_az=10*math.log10(islr)

        t=10*np.log10(t)
        
        mini=min(filter(lambda x:x!=float('-inf'),t))
        t[np.isnan(t)]=mini
        t[np.isinf(t)]=mini
        tmax=np.nanmax(t)
        t=t-tmax
        power_tm=spline(x0[0],t,xnew)
        
        
        minima=argrelextrema(power_tm, np.less)
        maxima=argrelextrema(power_tm, np.greater)
        
        
        mlsl=power_tm[maxima]
        
        
        Is=np.max(mlsl[mlsl<=0])
        
        pslr_az=Is
        
        
        maxi=power_tm[maxima]
        maxi2=maxima[0]
        
        aa=int(np.argwhere(maxi==np.max(maxi)))
        pos = (np.abs(power_tm-(-3))).argmin()
        
        a11=np.argsort(np.abs(power_tm[0:maxi2[aa]]-(-3)))[0]
        a22=np.argsort(np.abs(power_tm[maxi2[aa]:]-(-3)))[0]
        a22=maxi2[aa]+a22
        
        lin_space=d
        resolution_az=(abs(xnew[a11])+abs(xnew[a22]))*lin_space
        
        ax1.plot(x0[0],t,'go',xnew,power_tm,'b-',linewidth=2.0)

        ax1.set_xlabel('Azimuth (m)',fontsize=20)
        ax1.set_ylabel('Relative Power (dB)',fontsize=20) 
        ax1.tick_params(axis='x',labelsize=20)
        ax1.tick_params(axis='y',labelsize=20)
        ax1.axis('on')


        plt.show()     
          
          
   
    si="PSLR-Range : {:0.2f} dB".format(pslr_r)
    ssi="\nISLR-Range : {:0.2f} dB".format(islr_range)
    ti="\n\nResolution-Range : {:0.2f} m".format(resolution_r)
    ui="\nPSLR-Azimuth : {:0.2f} dB".format(pslr_az)
    uui="\nISLR-Azimuth : {:0.2f} dB".format(islr_az)
    vi="\nResolution-Azimuth : {:0.2f} m".format(resolution_az)
    tex1=Text(sar_top,height=10,width=50)
    tex1.grid()
    tex1.insert(END,si+ssi+ti+ui+uui+vi)
    tex1.mainloop()
        
         
         
    sar_top.mainloop()    
         
     

b7 = Button(root,text="SAR Image Quality Parameters", command=sar_qual)
print('\n\n\n')
b7.grid(row=900,column=0,sticky='W',padx=round(sc_w/2))



root.mainloop()






















